"""Settings + back-office endpoints for the in-POS settings page
(promotions, price books, delivery apps, discount approval)."""

import json

import frappe
from frappe import _
from frappe.utils import cint, flt, getdate, nowdate
from frappe.utils.password import get_decrypted_password

from lumenpos import __version__

def _can_manage():
    """Can the user change LumenPOS-wide settings (the General tab)?"""
    return bool(frappe.has_permission("LumenPOS Settings", "write"))


def _require(doctype, ptype="write"):
    """Enforce a standard ERPNext DocType permission at the API boundary, so
    every back-office action is governed by the Role Permissions Manager."""
    if not frappe.has_permission(doctype, ptype):
        frappe.throw(
            _("You are not permitted to {0} {1}.").format(_(ptype), _(doctype)),
            frappe.PermissionError,
        )


def _require_manager():
    _require("LumenPOS Settings", "write")


def _default_selling_price_list():
    """The price list a new price book starts on — the site's default selling
    price list, falling back to ERPNext's 'Standard Selling'."""
    return frappe.db.get_single_value("Selling Settings", "selling_price_list") or (
        "Standard Selling" if frappe.db.exists("Price List", "Standard Selling") else None
    )


def _protected_price_lists():
    """Base selling lists a price book / delivery app must NOT edit — editing
    them would change real selling prices, not create overrides. Covers the
    company's default selling list, every POS Profile's selling list and
    ERPNext's 'Standard Selling'."""
    names = set()
    ss = frappe.db.get_single_value("Selling Settings", "selling_price_list")
    if ss:
        names.add(ss)
    for pl in frappe.get_all("POS Profile", pluck="selling_price_list"):
        if pl:
            names.add(pl)
    if frappe.db.exists("Price List", "Standard Selling"):
        names.add("Standard Selling")
    return names


def _block_if_protected(price_list):
    if price_list and price_list in _protected_price_lists():
        frappe.throw(
            _(
                "“{0}” is a base selling price list — editing it here would change your "
                "normal selling prices. Create a dedicated price list for this book/app instead."
            ).format(price_list)
        )


@frappe.whitelist()
def get_settings():
    from lumenpos.api.session import get_user_permissions

    doc = frappe.get_cached_doc("LumenPOS Settings")
    return {
        "version": __version__,
        "can_manage": _can_manage(),
        "permissions": get_user_permissions(),
        "default_price_list": _default_selling_price_list(),
        "protected_price_lists": sorted(_protected_price_lists()),
        "allow_multiple_opening": doc.get("allow_multiple_opening") or 0,
        "offline_stock_only": doc.get("offline_stock_only") or 0,
        "show_out_of_stock": doc.get("show_out_of_stock") or 0,
        "serial_scan_only": 1 if doc.get("serial_scan_only") else 0,
        "gift_card_expiry_days": doc.get("gift_card_expiry_days") or 0,
        "gift_card_mode_of_payment": doc.get("gift_card_mode_of_payment") or "",
        "gift_card_account": doc.get("gift_card_account") or "",
        "gift_card_item": doc.get("gift_card_item") or "",
        "restrict_refund_to_paid_mode": 1 if doc.get("restrict_refund_to_paid_mode") else 0,
        "refund_rules": [
            {"paid_mode": r.paid_mode, "refund_mode": r.refund_mode}
            for r in (doc.get("refund_rules") or [])
        ],
        "return_reasons": [
            r.reason for r in (doc.get("return_reasons") or []) if (r.reason or "").strip()
        ],
        "discount_limit_percent": flt(doc.discount_limit_percent),
        "discount_approval_mode": doc.get("discount_approval_mode") or "Passcode only",
        "approver_role": doc.get("approver_role") or "",
        "restrict_returns_to_window": 1 if doc.get("restrict_returns_to_window") else 0,
        "return_window_days": cint(doc.get("return_window_days")) or 0,
        "has_passcode": bool(
            get_decrypted_password(
                "LumenPOS Settings", "LumenPOS Settings", "discount_passcode", raise_exception=False
            )
        ),
        "approvers": [
            {"approver_name": row.approver_name, "has_passcode": bool(row.passcode)}
            for row in (doc.get("discount_approvers") or [])
        ],
        "delivery_apps": [
            {
                "app_name": row.app_name,
                "price_list": row.price_list,
                "require_order_id": row.require_order_id,
            }
            for row in (doc.delivery_apps or [])
        ],
    }


@frappe.whitelist()
def save_settings(payload):
    _require_manager()
    if isinstance(payload, str):
        payload = json.loads(payload)
    doc = frappe.get_doc("LumenPOS Settings")
    doc.allow_multiple_opening = 1 if payload.get("allow_multiple_opening") else 0
    doc.offline_stock_only = 1 if payload.get("offline_stock_only") else 0
    doc.show_out_of_stock = 1 if payload.get("show_out_of_stock") else 0
    doc.serial_scan_only = 1 if payload.get("serial_scan_only") else 0
    doc.gift_card_expiry_days = int(payload.get("gift_card_expiry_days") or 0)
    doc.gift_card_mode_of_payment = payload.get("gift_card_mode_of_payment") or None
    doc.gift_card_account = payload.get("gift_card_account") or None
    doc.gift_card_item = payload.get("gift_card_item") or None
    doc.restrict_refund_to_paid_mode = 1 if payload.get("restrict_refund_to_paid_mode") else 0
    doc.refund_rules = []
    for row in payload.get("refund_rules") or []:
        if (row.get("paid_mode") or "").strip() and (row.get("refund_mode") or "").strip():
            doc.append(
                "refund_rules",
                {"paid_mode": row["paid_mode"], "refund_mode": row["refund_mode"]},
            )
    if "return_reasons" in payload:
        doc.return_reasons = []
        seen_reasons = set()
        for reason in payload.get("return_reasons") or []:
            reason = (reason or "").strip()
            key = reason.casefold()
            if reason and key not in seen_reasons:
                seen_reasons.add(key)
                doc.append("return_reasons", {"reason": reason})
    doc.discount_limit_percent = flt(payload.get("discount_limit_percent"))
    doc.discount_approval_mode = payload.get("discount_approval_mode") or "Passcode only"
    doc.approver_role = payload.get("approver_role") or None
    doc.restrict_returns_to_window = 1 if payload.get("restrict_returns_to_window") else 0
    doc.return_window_days = cint(payload.get("return_window_days")) or 0
    if payload.get("discount_passcode"):
        doc.discount_passcode = payload["discount_passcode"]

    if "approvers" in payload:
        existing_pins = {
            row.approver_name: row.passcode for row in (doc.get("discount_approvers") or [])
        }
        doc.discount_approvers = []
        for row in payload.get("approvers") or []:
            name = (row.get("approver_name") or "").strip()
            if not name:
                continue
            doc.append(
                "discount_approvers",
                {
                    "approver_name": name,
                    # blank PIN in the form keeps the stored one
                    "passcode": row.get("passcode") or existing_pins.get(name),
                },
            )
    doc.delivery_apps = []
    for row in payload.get("delivery_apps") or []:
        if not (row.get("app_name") or "").strip():
            continue
        doc.append(
            "delivery_apps",
            {
                "app_name": row["app_name"].strip(),
                "price_list": row.get("price_list") or None,
                "require_order_id": 1 if row.get("require_order_id") else 0,
            },
        )
    doc.save()
    return get_settings()


def check_passcode(passcode):
    """Returns the approver's name when the passcode matches one of the
    named approvers, True for the legacy master passcode, None otherwise."""
    if not passcode:
        return None
    stored = get_decrypted_password(
        "LumenPOS Settings", "LumenPOS Settings", "discount_passcode", raise_exception=False
    )
    if stored and passcode == stored:
        return True
    doc = frappe.get_cached_doc("LumenPOS Settings")
    for row in doc.get("discount_approvers") or []:
        pin = get_decrypted_password(
            "POS Discount Approver", row.name, "passcode", raise_exception=False
        )
        if pin and passcode == pin:
            return row.approver_name
    return None


@frappe.whitelist()
def verify_passcode(passcode):
    result = check_passcode(passcode)
    return {
        "valid": bool(result),
        "approver": result if isinstance(result, str) else None,
    }


# ---------------------------------------------------------------------------
# Promotions CRUD (simplified shape for the Vend-style editor)
# ---------------------------------------------------------------------------

PROMO_FIELDS = [
    "title", "status", "promotion_type", "description", "priority", "stackable", "price_basis",
    "start_date", "end_date", "start_time", "end_time",
    "monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday",
    "customer_eligibility", "apply_on_all",
    "discount_type", "discount_value",
    "buy_qty", "get_qty", "get_discount_type", "get_discount_value", "max_applications",
    "min_spend", "basket_discount_type", "basket_discount_value",
    "bundle_price", "requires_coupon", "coupon_code",
]


@frappe.whitelist()
def list_promotions():
    today = getdate(nowdate())
    rows = frappe.get_all(
        "POS Promotion",
        fields=["name", "title", "promotion_type", "status", "start_date", "end_date",
                "requires_coupon", "coupon_code"],
        order_by="modified desc",
        limit_page_length=200,
    )
    for row in rows:
        row["expired"] = bool(row.end_date and getdate(row.end_date) < today)
    return rows


@frappe.whitelist()
def get_promotion(name):
    doc = frappe.get_doc("POS Promotion", name)
    doc.check_permission("read")
    data = {f: doc.get(f) for f in PROMO_FIELDS}
    data["name"] = doc.name
    data["pos_profiles"] = [r.pos_profile for r in (doc.pos_profiles or [])]
    data["customer_groups"] = [r.customer_group for r in (doc.customer_groups or [])]
    data["items"] = [
        {
            "applies_to": r.applies_to,
            "item_code": r.item_code,
            "item_group": r.item_group,
            "brand": r.brand,
            "tag": r.get("tag"),
            "role": r.role,
            "qty": r.qty,
            "exclude": r.get("exclude") or 0,
        }
        for r in (doc.items or [])
    ]
    return data


@frappe.whitelist()
def save_promotion(payload):
    if isinstance(payload, str):
        payload = json.loads(payload)

    if payload.get("name"):
        _require("POS Promotion", "write")
        doc = frappe.get_doc("POS Promotion", payload["name"])
    else:
        _require("POS Promotion", "create")
        doc = frappe.new_doc("POS Promotion")

    # Empty date/time inputs arrive as "" — store None, never 00:00:00
    # (a 00:00-00:00 window would silently disable the promotion)
    for field in ("start_date", "end_date", "start_time", "end_time"):
        if field in payload and not payload[field]:
            payload[field] = None

    for field in PROMO_FIELDS:
        if field in payload:
            doc.set(field, payload[field])

    doc.pos_profiles = []
    for profile in payload.get("pos_profiles") or []:
        doc.append("pos_profiles", {"pos_profile": profile})
    doc.customer_groups = []
    for group in payload.get("customer_groups") or []:
        doc.append("customer_groups", {"customer_group": _resolve_link("Customer Group", group)})
    doc.items = []
    for row in payload.get("items") or []:
        applies_to = row.get("applies_to") or "Item"
        value = (
            row.get("item_code")
            or row.get("item_group")
            or row.get("brand")
            or row.get("tag")
        )
        resolved = _resolve_link(applies_to if applies_to != "Item" else "Item", value)
        doc.append(
            "items",
            {
                "applies_to": applies_to,
                "item_code": resolved if applies_to == "Item" else None,
                "item_group": resolved if applies_to == "Item Group" else None,
                "brand": resolved if applies_to == "Brand" else None,
                "tag": resolved if applies_to == "Tag" else None,
                "role": row.get("role") or "Buy",
                "qty": row.get("qty") or 1,
                "exclude": 1 if row.get("exclude") else 0,
            },
        )

    doc.save()
    return doc.name


def _resolve_link(doctype, value):
    """STRICT link resolution for the settings editor: the value must be a
    real record. For items, a pasted item NAME is resolved to its code —
    anything unknown is rejected with a clear message instead of saving a
    promotion that can never match."""
    value = (value or "").strip()
    if not value:
        frappe.throw(_("Pick a {0} from the list").format(_(doctype)))
    if frappe.db.exists(doctype, value):
        return value
    if doctype == "Item":
        by_name = frappe.db.get_value("Item", {"item_name": value}, "name")
        if by_name:
            return by_name
        by_barcode = frappe.db.get_value("Item Barcode", {"barcode": value}, "parent")
        if by_barcode:
            return by_barcode
    frappe.throw(
        _("{0} '{1}' does not exist — pick it from the dropdown list").format(_(doctype), value)
    )


@frappe.whitelist()
def delete_promotion(name):
    _require("POS Promotion", "delete")
    frappe.delete_doc("POS Promotion", name)


@frappe.whitelist()
def test_promotion(name, items, pos_profile, customer_group=None):
    """Dry-run a promotion against a test basket of REAL items, with a
    gate-by-gate diagnosis. Turns 'the promotion is not working' into a
    visible reason: wrong day, wrong outlet, rows matching nothing, items
    without a price, and so on."""
    from frappe.utils import now_datetime

    from lumenpos.api.sales import _build_lines
    from lumenpos.promotions import engine
    from lumenpos.promotions.loader import serialize

    if isinstance(items, str):
        items = json.loads(items)
    if not items:
        frappe.throw(_("Add at least one item to the test basket"))

    doc = frappe.get_doc("POS Promotion", name)
    promo = serialize(doc)
    profile = frappe.get_cached_doc("POS Profile", pos_profile)
    lines = _build_lines(
        [{"item_code": i["item_code"], "qty": i.get("qty") or 1} for i in items],
        profile,
    )
    now = now_datetime()
    cart = {
        "customer_group": customer_group,
        "pos_profile": pos_profile,
        # the tester auto-supplies the coupon so the offer itself can be tested
        "coupon_codes": [promo["coupon_code"]] if promo["requires_coupon"] else [],
        "items": lines,
    }

    date_str = now.strftime("%Y-%m-%d")
    day_field = engine.DAYS[now.weekday()]
    gates = [
        {"label": _("Status is Active"), "ok": promo["status"] == "Active"},
        {
            "label": _("Date window ({0} → {1})").format(
                promo["start_date"] or "…", promo["end_date"] or "…"
            ),
            "ok": (not promo["start_date"] or date_str >= promo["start_date"])
            and (not promo["end_date"] or date_str <= promo["end_date"]),
        },
        {
            "label": _("Runs on {0}").format(day_field.capitalize()),
            "ok": not any((promo["days"] or {}).values()) or bool(promo["days"].get(day_field)),
        },
        {
            "label": _("Daily time window"),
            "ok": engine.is_active(
                {**promo, "pos_profiles": [], "customer_eligibility": "All Customers",
                 "requires_coupon": 0, "days": {}, "start_date": None, "end_date": None},
                now, cart,
            ),
        },
        {
            "label": _("Outlet {0} included").format(pos_profile),
            "ok": not promo["pos_profiles"] or pos_profile in promo["pos_profiles"],
        },
        {
            "label": _("Customer group eligibility"),
            "ok": promo["customer_eligibility"] != "Selected Customer Groups"
            or (customer_group and customer_group in promo["customer_groups"]),
        },
    ]

    row_matches = []
    for row in promo["items"]:
        matching = [
            line["item_code"] for line in lines if engine._line_matches(line, [row])
        ]
        row_matches.append(
            {
                "applies_to": row["applies_to"],
                "value": row["value"],
                "role": row["role"],
                "matches": matching,
            }
        )

    price_warnings = [
        _("{0} has NO price on price list '{1}' — a percentage of 0 is 0").format(
            line["item_code"], profile.selling_price_list
        )
        for line in lines
        if not line["price"]
    ]

    result = engine.evaluate(cart, [promo], now)
    return {
        "gates": gates,
        "active": engine.is_active(promo, now, cart),
        "row_matches": row_matches,
        "lines": [
            {"item_code": line["item_code"], "qty": line["qty"], "price": line["price"]}
            for line in lines
        ],
        "price_warnings": price_warnings,
        "result": result,
    }


# ---------------------------------------------------------------------------
# Bundles CRUD
# ---------------------------------------------------------------------------

@frappe.whitelist()
def list_bundles():
    today = getdate(nowdate())
    bundles = frappe.get_all(
        "POS Bundle",
        fields=["name", "title", "bundle_price", "status", "valid_from", "valid_to"],
        order_by="modified desc",
        limit_page_length=200,
    )
    for bundle in bundles:
        doc = frappe.get_doc("POS Bundle", bundle.name)
        bundle["items"] = [
            {
                "item_code": row.item_code,
                "item_name": row.item_name,
                "qty": row.qty,
                "allocated_amount": row.get("allocated_amount"),
            }
            for row in doc.items
        ]
        bundle["pos_profiles"] = [row.pos_profile for row in (doc.pos_profiles or [])]
        bundle["expired"] = bool(bundle.valid_to and getdate(bundle.valid_to) < today)
    return bundles


@frappe.whitelist()
def save_bundle(payload):
    if isinstance(payload, str):
        payload = json.loads(payload)
    if payload.get("name"):
        _require("POS Bundle", "write")
        doc = frappe.get_doc("POS Bundle", payload["name"])
    else:
        _require("POS Bundle", "create")
        doc = frappe.new_doc("POS Bundle")
    doc.title = payload.get("title")
    doc.status = payload.get("status") or "Active"
    doc.bundle_price = payload.get("bundle_price")
    doc.valid_from = payload.get("valid_from") or None
    doc.valid_to = payload.get("valid_to") or None
    doc.items = []
    for row in payload.get("items") or []:
        doc.append(
            "items",
            {
                "item_code": _resolve_link("Item", row.get("item_code")),
                "qty": row.get("qty") or 1,
                "allocated_amount": flt(row.get("allocated_amount")) or None,
            },
        )
    doc.pos_profiles = []
    for profile in payload.get("pos_profiles") or []:
        doc.append("pos_profiles", {"pos_profile": profile})
    doc.save()
    return doc.name


@frappe.whitelist()
def delete_bundle(name):
    _require("POS Bundle", "delete")
    frappe.delete_doc("POS Bundle", name)


# ---------------------------------------------------------------------------
# Price books CRUD
# ---------------------------------------------------------------------------

@frappe.whitelist()
def list_price_books():
    today = getdate(nowdate())
    books = frappe.get_all(
        "POS Price Book",
        fields=["name", "title", "price_list", "status", "priority", "valid_from", "valid_to"],
        order_by="priority desc",
        limit_page_length=100,
    )
    for book in books:
        doc = frappe.get_cached_doc("POS Price Book", book.name)
        book["expired"] = bool(book.valid_to and getdate(book.valid_to) < today)
        book["pos_profiles"] = [r.pos_profile for r in (doc.pos_profiles or [])]
        book["customer_groups"] = [r.customer_group for r in (doc.customer_groups or [])]
        book["items"] = [
            {
                "item_code": r.item_code,
                "item_name": r.item_name
                or frappe.get_cached_value("Item", r.item_code, "item_name"),
                "rate": r.rate,
            }
            for r in (doc.get("items") or [])
        ]
    return books


@frappe.whitelist()
def save_price_book(payload):
    if isinstance(payload, str):
        payload = json.loads(payload)
    if payload.get("name"):
        _require("POS Price Book", "write")
        doc = frappe.get_doc("POS Price Book", payload["name"])
    else:
        _require("POS Price Book", "create")
        doc = frappe.new_doc("POS Price Book")
    for field in ("title", "status", "priority", "valid_from", "valid_to"):
        if field in payload:
            doc.set(field, payload[field] or None)
    doc.pos_profiles = []
    for profile in payload.get("pos_profiles") or []:
        doc.append("pos_profiles", {"pos_profile": profile})
    doc.customer_groups = []
    for group in payload.get("customer_groups") or []:
        doc.append("customer_groups", {"customer_group": group})
    doc.items = []
    for row in payload.get("items") or []:
        item_code = _resolve_item(row.get("item_code"))
        if not item_code:
            continue
        rate = flt(row.get("rate"))
        if rate < 0:
            frappe.throw(_("Price for {0} cannot be negative").format(item_code))
        doc.append("items", {"item_code": item_code, "rate": rate})
    doc.save()
    return doc.name


@frappe.whitelist()
def delete_price_book(name):
    _require("POS Price Book", "delete")
    frappe.delete_doc("POS Price Book", name)


@frappe.whitelist()
def list_price_book_prices(price_list, search="", start=0, limit=50, compare_price_list=None):
    """Items priced on a price book's price list, with the default price
    alongside for comparison."""
    conditions = "ip.price_list = %(price_list)s and ip.selling = 1"
    params = {"price_list": price_list, "start": int(start), "limit": min(int(limit), 100)}
    if search:
        conditions += " and (ip.item_code like %(search)s or i.item_name like %(search)s)"
        params["search"] = f"%{search}%"

    rows = frappe.db.sql(
        f"""
        select ip.item_code, i.item_name, ip.price_list_rate as rate
        from `tabItem Price` ip
        join `tabItem` i on i.name = ip.item_code
        where {conditions}
        order by i.item_name asc
        limit %(start)s, %(limit)s
        """,
        params,
        as_dict=True,
    )
    total = frappe.db.sql(
        f"""
        select count(*) from `tabItem Price` ip
        join `tabItem` i on i.name = ip.item_code
        where {conditions}
        """,
        params,
    )[0][0]

    if compare_price_list and rows:
        from lumenpos.price_books import get_price_map

        compare = get_price_map([r.item_code for r in rows], compare_price_list)
        for row in rows:
            row["default_rate"] = compare.get(row.item_code)
    return {"items": rows, "total": total}


def _upsert_item_price(price_list, item_code, rate):
    """Set one item's selling price on a price list; returns 'created' or
    'updated'. Shared by the inline editor and the bulk importer."""
    rate = flt(rate)
    if rate < 0:
        frappe.throw(_("Price cannot be negative"))
    existing = frappe.db.get_value(
        "Item Price",
        {"price_list": price_list, "item_code": item_code, "selling": 1},
        "name",
    )
    if existing:
        doc = frappe.get_doc("Item Price", existing)
        doc.price_list_rate = rate
        doc.flags.ignore_permissions = True
        doc.save()
        return "updated"
    frappe.get_doc(
        {
            "doctype": "Item Price",
            "item_code": item_code,
            "price_list": price_list,
            "price_list_rate": rate,
            "selling": 1,
        }
    ).insert(ignore_permissions=True)
    return "created"


def _resolve_item(value):
    """Item code from a code, an item name, or a barcode — or None."""
    value = (value or "").strip()
    if not value:
        return None
    if frappe.db.exists("Item", value):
        return value
    by_name = frappe.db.get_value("Item", {"item_name": value}, "name")
    if by_name:
        return by_name
    return frappe.db.get_value("Item Barcode", {"barcode": value}, "parent")


@frappe.whitelist()
def set_price_book_price(price_list, item_code, rate):
    """Upsert one item's price on a price book's price list."""
    _require("Item Price", "write")
    _block_if_protected(price_list)
    _upsert_item_price(price_list, item_code, rate)
    return {"item_code": item_code, "rate": flt(rate)}


@frappe.whitelist()
def export_price_book_prices(price_list, compare_price_list=None):
    """Download every item priced on this price list as a CSV the user can
    edit in Excel and re-import. Columns: Item Code, Item Name, Barcode,
    Default Price, Book Price."""
    _require("Item Price", "read")
    import csv as csvmod
    import io

    rows = frappe.db.sql(
        """
        select ip.item_code, i.item_name, ip.price_list_rate as rate
        from `tabItem Price` ip
        join `tabItem` i on i.name = ip.item_code
        where ip.price_list = %(pl)s and ip.selling = 1
        order by i.item_name asc
        """,
        {"pl": price_list},
        as_dict=True,
    )
    codes = [r.item_code for r in rows]
    barcode_map = {}
    if codes:
        for b in frappe.get_all(
            "Item Barcode",
            filters={"parent": ["in", codes]},
            fields=["parent", "barcode"],
            order_by="idx asc",
        ):
            barcode_map.setdefault(b.parent, b.barcode)
    compare = {}
    if compare_price_list and compare_price_list != price_list and codes:
        from lumenpos.price_books import get_price_map

        compare = get_price_map(codes, compare_price_list)

    buf = io.StringIO()
    writer = csvmod.writer(buf)
    writer.writerow(["Item Code", "Item Name", "Barcode", "Default Price", "Book Price"])
    for r in rows:
        writer.writerow(
            [
                r.item_code,
                r.item_name,
                barcode_map.get(r.item_code, ""),
                compare.get(r.item_code, ""),
                r.rate,
            ]
        )
    return {"filename": f"{price_list} - prices.csv", "csv": buf.getvalue()}


@frappe.whitelist()
def import_price_book_prices(price_list, filename, content):
    """Bulk-set prices from an uploaded .xlsx or .csv file. Rows are matched
    to items by code, name or barcode; the price comes from a Book Price /
    Price / Rate column (or the last column if there's no header). Returns a
    created/updated count plus any rows that couldn't be matched."""
    _require("Item Price", "write")
    _block_if_protected(price_list)
    if not frappe.db.exists("Price List", price_list):
        frappe.throw(_("Price list {0} does not exist").format(price_list))

    rows = _parse_price_table(filename, content)
    updated = created = 0
    errors = []
    for row in rows:
        if row.get("error"):
            errors.append(row["error"])
            continue
        result = _upsert_item_price(price_list, row["item_code"], row["rate"])
        created += result == "created"
        updated += result == "updated"
    return {"updated": updated, "created": created, "errors": errors[:50]}


def _parse_price_table(filename, content):
    """Decode an uploaded .xlsx/.csv of (item, price) rows and match each row to
    an item by code, name or barcode. Returns dicts {item_code, item_name,
    rate} for good rows and {error} for bad ones, in file order."""
    import base64
    import io

    raw = content.split(",", 1)[-1] if "," in content[:100] else content
    data = base64.b64decode(raw)
    ext = (filename or "").rsplit(".", 1)[-1].lower()

    table = []
    if ext in ("xlsx", "xlsm", "xls"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        for row in wb.active.iter_rows(values_only=True):
            table.append(list(row))
    else:
        import csv as csvmod

        text = data.decode("utf-8-sig", errors="ignore")
        table = [row for row in csvmod.reader(io.StringIO(text))]

    table = [r for r in table if any(c is not None and str(c).strip() for c in r)]
    if not table:
        frappe.throw(_("The file has no rows"))

    header = [str(c or "").strip().lower() for c in table[0]]

    def find(names):
        for idx, h in enumerate(header):
            if h in names:
                return idx
        return None

    code_idx = find({"item code", "item_code", "code", "item", "itemcode"})
    name_idx = find({"item name", "item_name", "name"})
    barcode_idx = find({"barcode", "ean", "upc", "bar code"})
    price_idx = find(
        {"book price", "price", "rate", "new price", "price_list_rate",
         "book_price", "amount", "selling price"}
    )
    has_header = any(i is not None for i in (code_idx, price_idx, barcode_idx))
    if has_header:
        body = table[1:]
    else:
        body, code_idx, price_idx = table, 0, len(header) - 1

    def cell(row, i):
        if i is None or i >= len(row) or row[i] is None:
            return ""
        return str(row[i]).strip()

    out = []
    for offset, row in enumerate(body):
        line_no = offset + (2 if has_header else 1)
        raw_code = cell(row, code_idx)
        price_raw = cell(row, price_idx)
        if not raw_code and not price_raw and not cell(row, barcode_idx):
            continue
        item_code = _resolve_item(raw_code)
        if not item_code and barcode_idx is not None:
            item_code = frappe.db.get_value(
                "Item Barcode", {"barcode": cell(row, barcode_idx)}, "parent"
            )
        if not item_code and name_idx is not None:
            item_code = _resolve_item(cell(row, name_idx))
        if not item_code:
            out.append({"error": _("Row {0}: item '{1}' not found").format(
                line_no, raw_code or cell(row, barcode_idx) or cell(row, name_idx))})
            continue
        if not price_raw:
            out.append({"error": _("Row {0}: no price for {1}").format(line_no, item_code)})
            continue
        try:
            rate = flt(price_raw.replace(",", ""))
        except Exception:
            out.append({"error": _("Row {0}: invalid price '{1}'").format(line_no, price_raw)})
            continue
        out.append({
            "item_code": item_code,
            "item_name": frappe.get_cached_value("Item", item_code, "item_name"),
            "rate": rate,
        })
    return out


def _item_group_descendants(group):
    """An item group and all its sub-groups (nested set), so 'Electronics'
    pulls everything under it."""
    node = frappe.db.get_value("Item Group", group, ["lft", "rgt"], as_dict=True)
    if not node:
        return [group]
    return frappe.get_all(
        "Item Group", filters={"lft": [">=", node.lft], "rgt": ["<=", node.rgt]}, pluck="name"
    )


def _items_with_tag(tag):
    """Item codes carrying an ERPNext tag — via the Tag Link table, falling back
    to the denormalised _user_tags string."""
    codes = frappe.get_all(
        "Tag Link",
        filters={"document_type": "Item", "tag": tag},
        pluck="document_name",
    )
    if codes:
        return codes
    return frappe.get_all("Item", filters={"_user_tags": ["like", f"%{tag}%"]}, pluck="name")


@frappe.whitelist()
def resolve_items(brand=None, item_group=None, tag=None):
    """Sellable item codes matching a brand / item group (incl. sub-groups) /
    tag — for bulk-adding to a price book in one click. Returns
    [{item_code, item_name}]."""
    _require("Item", "read")
    brand = (brand or "").strip()
    item_group = (item_group or "").strip()
    tag = (tag or "").strip()
    if not (brand or item_group or tag):
        return []
    filters = {"disabled": 0, "is_sales_item": 1}
    if brand:
        filters["brand"] = brand
    if item_group:
        groups = _item_group_descendants(item_group)
        filters["item_group"] = ["in", groups]
    if tag:
        tagged = _items_with_tag(tag)
        if not tagged:
            return []
        filters["name"] = ["in", tagged]
    items = frappe.get_all(
        "Item",
        filters=filters,
        fields=["name as item_code", "item_name"],
        order_by="item_name asc",
        limit_page_length=0,
    )
    # Default each row to its current selling price so a bulk add is non-
    # destructive (an item added at 0 would otherwise sell free); the manager
    # then lowers the ones they want to discount.
    selling_list = frappe.db.get_single_value("Selling Settings", "selling_price_list")
    price_map = {}
    if selling_list and items:
        for p in frappe.get_all(
            "Item Price",
            filters={
                "price_list": selling_list,
                "selling": 1,
                "item_code": ["in", [i.item_code for i in items]],
            },
            fields=["item_code", "price_list_rate"],
        ):
            price_map.setdefault(p.item_code, p.price_list_rate)
    for i in items:
        i["rate"] = flt(price_map.get(i.item_code) or 0)
    return items


# ---------------------------------------------------------------------------
# Coupon pool (bulk codes per promotion)
# ---------------------------------------------------------------------------

@frappe.whitelist()
def generate_coupons(promotion, count, prefix=None, usage_limit=1, valid_until=None):
    """Create `count` unique random codes for a coupon-locked promotion.
    usage_limit = redemptions allowed per code (0 = unlimited)."""
    _require("POS Coupon", "create")
    from lumenpos import coupons

    created = coupons.generate(
        promotion, count, prefix or "", cint(usage_limit), valid_until or None
    )
    return {"created": created, **coupons.summary(promotion)}


@frappe.whitelist()
def import_coupons(promotion, filename, content, usage_limit=1, valid_until=None):
    """Create coupons from an uploaded .xlsx/.csv of codes (first column)."""
    _require("POS Coupon", "create")
    from lumenpos import coupons

    codes = coupons.parse_codes_file(filename, content)
    if not codes:
        frappe.throw(_("No codes found in the file"))
    result = coupons.import_codes(promotion, codes, cint(usage_limit), valid_until or None)
    return {**result, **coupons.summary(promotion)}


@frappe.whitelist()
def list_coupons(promotion):
    """Coupon counts for a promotion (total / used / available)."""
    _require("POS Coupon", "read")
    from lumenpos import coupons

    return coupons.summary(promotion)


@frappe.whitelist()
def export_coupons(promotion):
    """All codes for a promotion, for CSV download / printing."""
    _require("POS Coupon", "read")
    from lumenpos import coupons

    return coupons.list_codes(promotion)


@frappe.whitelist()
def delete_coupons(promotion, only_unused=1):
    """Remove a promotion's coupons (unused only by default)."""
    _require("POS Coupon", "delete")
    filters = {"promotion": promotion}
    if cint(only_unused):
        filters["used"] = 0
    for name in frappe.get_all("POS Coupon", filters=filters, pluck="name"):
        frappe.delete_doc("POS Coupon", name)
    from lumenpos import coupons

    return coupons.summary(promotion)


@frappe.whitelist()
def parse_price_rows(filename, content):
    """Parse an uploaded .xlsx/.csv into item+price rows for the price-book item
    editor. The rows are merged into the book and saved with it — nothing is
    written to Item Price here."""
    _require("POS Price Book", "write")
    rows = _parse_price_table(filename, content)
    return {
        "items": [r for r in rows if not r.get("error")],
        "errors": [r["error"] for r in rows if r.get("error")][:50],
    }


@frappe.whitelist()
def remove_price_book_price(price_list, item_code):
    _require("Item Price", "delete")
    _block_if_protected(price_list)
    existing = frappe.db.get_value(
        "Item Price",
        {"price_list": price_list, "item_code": item_code, "selling": 1},
        "name",
    )
    if existing:
        frappe.delete_doc("Item Price", existing, ignore_permissions=True)


@frappe.whitelist()
def create_price_list(price_list_name):
    """Create a fresh selling Price List from the POS settings page so price
    books never require a trip to the desk."""
    _require("Price List", "create")
    price_list_name = (price_list_name or "").strip()
    if not price_list_name:
        frappe.throw(_("Enter a name for the price list"))
    if frappe.db.exists("Price List", price_list_name):
        return price_list_name
    currency = frappe.db.get_default("currency")
    if not currency:
        company = frappe.get_all("Company", fields=["default_currency"], limit=1)
        currency = company[0].default_currency if company else None
    doc = frappe.get_doc(
        {
            "doctype": "Price List",
            "price_list_name": price_list_name,
            "selling": 1,
            "enabled": 1,
            "currency": currency,
        }
    ).insert(ignore_permissions=True)
    return doc.name


# ---------------------------------------------------------------------------
# Loyalty
# ---------------------------------------------------------------------------

@frappe.whitelist()
def list_loyalty_programs():
    programs = frappe.get_all(
        "Loyalty Program",
        fields=[
            "name", "loyalty_program_name", "conversion_factor",
            "expiry_duration", "auto_opt_in", "company",
        ],
        limit_page_length=50,
    )
    for program in programs:
        rules = frappe.get_all(
            "Loyalty Program Collection",
            filters={"parent": program.name},
            fields=["collection_factor"],
            order_by="idx asc",
            limit_page_length=1,
        )
        program["collection_factor"] = rules[0].collection_factor if rules else 0
    return programs


@frappe.whitelist()
def create_loyalty_program(payload):
    """Simple single-tier program: earn 1 point per {spend_per_point} spent,
    each point worth {point_value} at redemption. Auto opt-in enrolls every
    customer."""
    _require("Loyalty Program", "create")
    if isinstance(payload, str):
        payload = json.loads(payload)

    name = (payload.get("name") or "").strip()
    if not name:
        frappe.throw(_("Give the loyalty program a name"))
    company = payload.get("company")
    spend_per_point = flt(payload.get("spend_per_point"))
    point_value = flt(payload.get("point_value"))
    if spend_per_point <= 0 or point_value <= 0:
        frappe.throw(_("Earning and redemption rates must be greater than zero"))

    expense_account = payload.get("expense_account") or _loyalty_expense_account(company)

    doc = frappe.get_doc(
        {
            "doctype": "Loyalty Program",
            "loyalty_program_name": name,
            "loyalty_program_type": "Single Tier Program",
            "company": company,
            "auto_opt_in": 1,
            "conversion_factor": point_value,
            "expiry_duration": int(payload.get("expiry_days") or 0),
            "expense_account": expense_account,
            "cost_center": frappe.get_cached_value("Company", company, "cost_center"),
            "collection_rules": [
                {"tier_name": "Standard", "collection_factor": spend_per_point, "min_spent": 0}
            ],
        }
    ).insert(ignore_permissions=True)
    return doc.name


def _loyalty_expense_account(company):
    abbr = frappe.get_cached_value("Company", company, "abbr")
    account_name = f"Loyalty Points Expense - {abbr}"
    if frappe.db.exists("Account", account_name):
        return account_name
    parent = frappe.db.get_value(
        "Account",
        {"company": company, "root_type": "Expense", "is_group": 1,
         "account_name": ["in", ["Indirect Expenses", "Expenses"]]},
        "name",
    ) or frappe.db.get_value(
        "Account", {"company": company, "root_type": "Expense", "is_group": 1}, "name"
    )
    if not parent:
        frappe.throw(_("No expense account group found for {0}").format(company))
    return (
        frappe.get_doc(
            {
                "doctype": "Account",
                "account_name": "Loyalty Points Expense",
                "parent_account": parent,
                "company": company,
                "root_type": "Expense",
            }
        )
        .insert(ignore_permissions=True)
        .name
    )


# ---------------------------------------------------------------------------
# Gift cards
# ---------------------------------------------------------------------------

@frappe.whitelist()
def list_gift_cards(search="", limit=50):
    filters = {}
    or_filters = None
    if search:
        or_filters = {
            "card_no": ["like", f"%{search}%"],
            "customer": ["like", f"%{search}%"],
        }
    return frappe.get_all(
        "POS Gift Card",
        filters=filters,
        or_filters=or_filters,
        fields=["card_no", "status", "initial_amount", "balance", "expiry_date", "customer"],
        order_by="creation desc",
        limit_page_length=min(int(limit), 100),
    )


@frappe.whitelist()
def disable_gift_card(card_no):
    _require("POS Gift Card", "write")
    frappe.db.set_value("POS Gift Card", card_no, "status", "Disabled")


# ---------------------------------------------------------------------------
# Pickers for the settings editor
# ---------------------------------------------------------------------------

@frappe.whitelist()
def link_options(doctype, search=""):
    allowed = {
        "Item": ["name", "item_name"],
        "Item Group": ["name"],
        "Brand": ["name"],
        "Customer Group": ["name"],
        "POS Profile": ["name"],
        "Price List": ["name"],
        "Account": ["name"],
        "Mode of Payment": ["name"],
        "Role": ["name"],
    }
    if doctype not in allowed:
        frappe.throw(_("Lookup not allowed for {0}").format(doctype))
    if doctype == "Item":
        return _item_link_options(search)
    or_filters = None
    if search:
        or_filters = {f: ["like", f"%{search}%"] for f in allowed[doctype]}
    filters = {}
    if doctype == "Price List":
        filters["selling"] = 1
    if doctype in ("Item Group", "Customer Group"):
        filters["is_group"] = 0
    if doctype == "Account":
        # Any postable account (loyalty wants Expense, gift cards want a
        # Liability account) — the field label guides the choice.
        filters["is_group"] = 0
    if doctype == "Mode of Payment":
        filters["enabled"] = 1
    if doctype == "Role":
        filters["disabled"] = 0
    results = frappe.get_all(
        doctype,
        filters=filters,
        or_filters=or_filters,
        fields=allowed[doctype],
        order_by="name asc",
        limit_page_length=20,
    )
    if doctype == "Price List":
        # Don't offer base selling lists as a price-book / app list — editing
        # them would change normal selling prices.
        protected = _protected_price_lists()
        results = [r for r in results if r["name"] not in protected]
    return results


def _item_link_options(search):
    """Item autocomplete for the settings editors — matches on item code,
    item name OR barcode, and returns the first barcode for display."""
    search = (search or "").strip()
    or_filters = None
    if search:
        or_filters = [
            ["name", "like", f"%{search}%"],
            ["item_name", "like", f"%{search}%"],
        ]
        codes = frappe.get_all(
            "Item Barcode",
            filters={"barcode": ["like", f"%{search}%"]},
            pluck="parent",
            limit_page_length=20,
        )
        if codes:
            or_filters.append(["name", "in", list(set(codes))])
    items = frappe.get_all(
        "Item",
        filters={"disabled": 0, "has_variants": 0, "is_sales_item": 1},
        or_filters=or_filters,
        fields=["name", "item_name"],
        order_by="item_name asc",
        limit_page_length=20,
    )
    if items:
        barcode_map = {}
        for row in frappe.get_all(
            "Item Barcode",
            filters={"parent": ["in", [i.name for i in items]]},
            fields=["parent", "barcode"],
            order_by="idx asc",
        ):
            barcode_map.setdefault(row.parent, row.barcode)
        for item in items:
            item["barcode"] = barcode_map.get(item.name)
    return items
