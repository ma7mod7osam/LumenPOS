"""Bulk coupon pool for LumenPOS promotions.

A coupon-locked POS Promotion can have MANY codes (generated or imported), each
unlocking that promotion's discount at the till. Codes are either single-use
(spent after one redemption) or reusable (work until they expire) — chosen per
batch. Legacy single `coupon_code` on the promotion keeps working alongside the
pool.
"""

import secrets

import frappe
from frappe import _
from frappe.utils import getdate, now_datetime, nowdate

# Human-friendly alphabet — no 0/O, 1/I/L to avoid mis-reads.
_ALPHABET = "ABCDEFGHJKMNPQRSTUVWXYZ23456789"


def _random_code(prefix="", length=8):
    body = "".join(secrets.choice(_ALPHABET) for _ in range(length))
    return f"{(prefix or '').strip().upper()}{body}"


def resolve(code):
    """A USABLE coupon row for `code`, or None. Usable = exists, not past its
    valid_until, and under its use limit (usage_limit 0 = unlimited). The doc
    name IS the code (autoname field:code)."""
    code = (code or "").strip()
    if not code:
        return None
    row = frappe.db.get_value(
        "POS Coupon",
        code,
        ["name", "code", "promotion", "usage_limit", "used_count", "valid_until"],
        as_dict=True,
    )
    if not row:
        return None
    limit = int(row.usage_limit or 0)
    if limit and int(row.used_count or 0) >= limit:
        return None
    if row.valid_until and getdate(row.valid_until) < getdate(nowdate()):
        return None
    return row


def apply_to_promotions(promotions, codes):
    """Mutate a serialized promo list so a valid pool code unlocks its promotion
    in the (DB-free) engine — by setting that promo's coupon_code to the code,
    exactly like a legacy single coupon. Returns the same list."""
    if not codes:
        return promotions
    by_name = {p.get("name"): p for p in promotions if p.get("name")}
    for code in codes:
        row = resolve(code)
        if row and row.promotion in by_name:
            by_name[row.promotion]["requires_coupon"] = 1
            by_name[row.promotion]["coupon_code"] = code
    return promotions


def consume(codes, invoice):
    """Count one redemption against each valid pool code used on `invoice`, and
    flag it Fully Used once its limit is reached. Unlimited codes (usage_limit 0)
    just accrue a count. No-op for unknown/expired codes. Safe to call with the
    raw cart codes."""
    for code in codes or []:
        row = resolve(code)
        if not row:
            continue
        new_count = int(row.used_count or 0) + 1
        limit = int(row.usage_limit or 0)
        frappe.db.set_value(
            "POS Coupon",
            row.name,
            {
                "used_count": new_count,
                "used": 1 if (limit and new_count >= limit) else 0,
                "used_on": invoice,
                "used_at": now_datetime(),
            },
        )


def generate(promotion, count, prefix="", usage_limit=1, valid_until=None, batch=None):
    """Create `count` unique random codes for a promotion. usage_limit is how
    many times EACH code may be redeemed (0 = unlimited). Returns the count
    actually created."""
    count = int(count or 0)
    if count <= 0:
        frappe.throw(_("How many coupons should I generate?"))
    if count > 5000:
        frappe.throw(_("Generate at most 5000 coupons at a time"))
    if not frappe.db.exists("POS Promotion", promotion):
        frappe.throw(_("Promotion {0} does not exist").format(promotion))
    batch = batch or _("Generated {0}").format(now_datetime().strftime("%Y-%m-%d %H:%M"))
    created = 0
    for _i in range(count):
        code = None
        for _attempt in range(6):
            candidate = _random_code(prefix)
            if not frappe.db.exists("POS Coupon", candidate):
                code = candidate
                break
        if not code:
            continue
        _insert(code, promotion, usage_limit, valid_until, batch)
        created += 1
    return created


def import_codes(promotion, codes, usage_limit=1, valid_until=None, batch=None):
    """Create coupons from a list of code strings (deduped; existing ones
    skipped). Returns {created, skipped}."""
    if not frappe.db.exists("POS Promotion", promotion):
        frappe.throw(_("Promotion {0} does not exist").format(promotion))
    batch = batch or _("Imported {0}").format(now_datetime().strftime("%Y-%m-%d %H:%M"))
    created, skipped, seen = 0, 0, set()
    for raw in codes or []:
        code = (raw or "").strip()
        key = code.upper()
        if not code or key in seen:
            continue
        seen.add(key)
        if frappe.db.exists("POS Coupon", code):
            skipped += 1
            continue
        _insert(code, promotion, usage_limit, valid_until, batch)
        created += 1
    return {"created": created, "skipped": skipped}


def _insert(code, promotion, usage_limit, valid_until, batch):
    limit = int(usage_limit or 0)
    frappe.get_doc(
        {
            "doctype": "POS Coupon",
            "code": code,
            "promotion": promotion,
            "usage_limit": limit,
            "used_count": 0,
            "single_use": 1 if limit == 1 else 0,  # legacy mirror
            "valid_until": valid_until or None,
            "batch": batch,
        }
    ).insert()


def summary(promotion):
    """Counts for the settings UI."""
    total = frappe.db.count("POS Coupon", {"promotion": promotion})
    used = frappe.db.count("POS Coupon", {"promotion": promotion, "used": 1})
    redemptions = (
        frappe.db.sql(
            "select coalesce(sum(used_count), 0) from `tabPOS Coupon` where promotion = %s",
            promotion,
        )[0][0]
        or 0
    )
    return {
        "total": total,
        "used": used,
        "available": total - used,
        "redemptions": int(redemptions),
    }


def list_codes(promotion, only_unused=False):
    """All codes for a promotion (for CSV export / printing)."""
    filters = {"promotion": promotion}
    if only_unused:
        filters["used"] = 0
    return frappe.get_all(
        "POS Coupon",
        filters=filters,
        fields=["code", "usage_limit", "used_count", "used", "valid_until", "batch"],
        order_by="creation desc",
        limit_page_length=0,
    )


def parse_codes_file(filename, content):
    """Extract a flat list of codes from an uploaded .xlsx/.csv — the first
    non-empty cell of each row (a 'code'/'coupon' header row is skipped)."""
    import base64
    import io

    raw = content.split(",", 1)[-1] if "," in content[:100] else content
    data = base64.b64decode(raw)
    ext = (filename or "").rsplit(".", 1)[-1].lower()

    rows = []
    if ext in ("xlsx", "xlsm", "xls"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        for row in wb.active.iter_rows(values_only=True):
            rows.append(list(row))
    else:
        import csv as csvmod

        text = data.decode("utf-8-sig", errors="ignore")
        rows = [r for r in csvmod.reader(io.StringIO(text))]

    codes = []
    for r in rows:
        first = next((str(c).strip() for c in r if c is not None and str(c).strip()), "")
        if first:
            codes.append(first)
    # Drop a header cell like "code" / "coupon".
    if codes and codes[0].lower() in ("code", "coupon", "coupon code", "coupons"):
        codes = codes[1:]
    return codes
